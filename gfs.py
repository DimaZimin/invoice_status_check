import glob
from openpyxl import load_workbook
from datetime import datetime

gfis_data = {}


def remove_row():
    for file in glob.glob('gfis/*.xlsx'):
        gfis_excel = load_workbook(filename=f'{file}')
        gfis_sheet = gfis_excel.active
        for col in gfis_sheet.iter_rows(max_row=1, values_only=True):
            print('Checking GFIS files...')
            if None in col:
                print(f'File: {file} - 1st row has been deleted')
                gfis_sheet.delete_rows(1)
                gfis_excel.save(file)
            else:
                print(f'No changes in {file}')
    print('GFIS files are OK.')


def pdate_column(file):
    gfis_excel = load_workbook(filename=f'{file}')
    gfis_sheet = gfis_excel.active
    for col in gfis_sheet.iter_rows(max_row=1):
        return len(col)


def retrieve_invoice_data():
    for file in glob.glob('gfis/*.xlsx'):
        try:
            gfis_excel = load_workbook(filename=f'{file}')
            gfis_sheet = gfis_excel.active
        except FileNotFoundError:
            print('File *.xlsx in <gfis> not found')
        else:
            gfis_invoices = [str(invoice[0]) for invoice in
                             gfis_sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True)]
            gfis_schedule = [datetime.strftime(schedule[0], '%Y-%m-%d') if schedule[0] else 'no data'
                             for schedule in gfis_sheet.iter_rows(min_row=2, min_col=13, max_col=13, values_only=True)]
            gfis_payment = [pdate[0] for pdate in gfis_sheet.iter_rows(min_row=2, min_col=pdate_column(file),
                                                                       max_col=pdate_column(file), values_only=True)]
            pmt_dates = [datetime.strftime(dt, '%Y-%m-%d') if dt is not None else 'not paid' for dt in gfis_payment]

            pmt = [pm[0] for pm in
                   gfis_sheet.iter_rows(min_row=2, min_col=9, max_col=9, values_only=True)]

            for inv, sch, dat, pm in zip(gfis_invoices, gfis_schedule, pmt_dates, pmt):
                if inv not in gfis_data.keys():
                    gfis_data[inv] = sch, dat, pm
                else:
                    if pm > gfis_data.get(inv)[2]:
                        gfis_data[inv] = sch, dat, pm
