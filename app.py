from bswr import decode_basware_files, combine_to_excel
from gfs import remove_row, retrieve_invoice_data, gfis_data
from flw import decoding_flow, combine_to_excel_flow
import glob
import openpyxl
from openpyxl import load_workbook
import os
from typing import List


bw_statuses = {'R1': 'Returned', 'C1': 'Cancelled',
               '20': 'Unprocessed E-invoice', '4': 'Cancelled Invoices',
               '3': 'Transferred to GFIS', '2': 'Ready to transfer',
               '1': 'Sent for approval', '0': 'Unprocessed'}

inv_basware_status = {}


def get_approver(invoice):
    for i, el in enumerate(flow_inv):
        if str(invoice) == str(el):
            return approvers[i], date_sent[i]


def inv_status():
    for inv in requested_invoices:
        if inv in gfis_data.keys():
            inv_basware_status[inv] = f'Scheduled due {gfis_data[inv][0]}, payment: {gfis_data[inv][1]}'
        elif inv in all_inv.keys():
            inv_basware_status[inv] = bw_statuses[str(all_inv[inv])]
        else:
            inv_basware_status[inv] = 'Missing'


def write_status(invoice_file):
    for i, (k, v) in enumerate(inv_basware_status.items()):
        invoice_sheet = invoice_file.active
        try:
            if v == bw_statuses['1']:
                invoice_sheet[f'B{i + 2}'] = f'{v} to {get_approver(k)[0]} on {get_approver(k)[1]}'
            elif v == bw_statuses['3']:
                invoice_sheet[f'B{i + 2}'] = f'{v}'
                invoice_sheet[f'C{i + 2}'] = f'NO DATA IN GFIS'
            else:
                invoice_sheet[f'B{i + 2}'] = f'{v}'
            invoice_file.save(filename='check_invoices.xlsx')
        except TypeError and KeyError:
            invoice_sheet[f'B{i + 2}'] = f'{v}'
            invoice_sheet[f'C{i + 2}'] = 'Data is missing. Please check manually. '


def remove_temporary_files():
    try:
        os.remove("all_invoices.xlsx")
        os.remove("flow/flow_invoices.xlsx")
    except FileNotFoundError:
        print('Nothing to remove.')


def load_invoices(invoice_file: openpyxl.workbook.Workbook) -> List[str]:
    """
    Loads first column in first sheet into a list.
    """
    invoice_sheet = invoice_file.active
    return [
        str(row[0]) for row in
        invoice_sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True)
    ]


if __name__ == '__main__':
    print('################### Basware/GFIS invoice status checking.##########################')
    print('============== Copyright © 2020 Dmytro Zimin. All Rights Reserved. ============')
    print('################################################################################\n\n')
    print('***Please make sure you have read the instruction manual before using this program***')
    run = input('To proceed press <y> or any other button to exit\n')
    if run.lower()[0] == 'y':
        try:
            invoice_file = load_workbook(filename='check_invoices.xlsx')
        except FileNotFoundError:
            print('Unable to find <check_invoices.xlsx> in root directory')
            input('Press any key to exit')
            exit()
        else:
            requested_invoices = load_invoices(invoice_file)

        print('Merging files...')
        combine_to_excel(input_directory='basware', output_file='all_invoices.xlsx')
        print('Merging files completed!')
        remove_row()
        print('Rewriting <flow> to .xlsx')
        combine_to_excel(input_directory='flow', output_file='flow/all_invoices.xlsx')
        try:
            print('Retrieving invoice data. Please wait...')
            path_file_flow = glob.glob('flow/*.xlsx')[0]
            inflow_invoices = load_workbook(filename=f'{path_file_flow}')
            inflow_sheet = inflow_invoices.active
        except FileNotFoundError:
            print('Error code 0. Please check if there is a *.xlxs file in <flow> directory')
            input('Press any key to exit')
            exit()
        else:
            approvers = [str(approver[0]) for approver in
                         inflow_sheet.iter_rows(min_row=2, min_col=13, max_col=13, values_only=True)]  # TODO: wrap into function, get rid of magic constants
            flow_inv = [str(inv[0]) for inv in
                        inflow_sheet.iter_rows(min_row=2, min_col=7, max_col=7, values_only=True)]
            date_sent = [dt[0].split(' ')[0] for dt in
                         inflow_sheet.iter_rows(min_row=2, min_col=15, max_col=15, values_only=True)]
        retrieve_invoice_data()
        try:
            all_invoices = load_workbook(filename='all_invoices.xlsx')
            allinv_sheet = all_invoices.active
        except FileNotFoundError:
            print('Cannot find <check_invoices.xlsx> file. Please read instructions and try again.')
            input('Press any key to exit')
            exit()
        else:
            all_inv_nr = [str(invoice[0]) for invoice in
                          allinv_sheet.iter_rows(min_row=2, min_col=9, max_col=9, values_only=True)]
            all_inv_st = [str(stat[0]) for stat in
                          allinv_sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True)]
            all_inv = {key: value for (key, value) in zip(all_inv_nr, all_inv_st)}
            inv_status()
            print('Updating <check_invoices.xlsx>...')
            write_status(invoice_file)
            print('Updated!')
            print('Removing temporary files...')
            remove_temporary_files()
            print('Removed!')
            input('Statuses have been added to check_invoices.xlsx! Press any key to exit')
            exit()
