import glob
import openpyxl
from openpyxl import load_workbook
import os
from typing import List


class InvoiceDataCheker:

    basware_csv_path = 'check_invoices.xlsx'

    def __init__(self, invoice_number, status):
        self.invoice_number = invoice_number
        self.status = status



def load_invoices_to_check(invoice_file: openpyxl.load_workbook(filename=InvoiceDataCheker.basware_csv_path)) -> List[str]:
    """
    Loads first column in first sheet into a list.
    """
    invoice_sheet = invoice_file.active
    return [
        str(row[0]) for row in
        invoice_sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True)
    ]


print(load_invoices_to_check(InvoiceDataCheker.basware_csv_path))
