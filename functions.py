from datafile import DataFile
import glob
from openpyxl import load_workbook
from mappings import *
from datetime import datetime


STATUS_CODES = {'R1': 'Returned', 'C1': 'Cancelled',
               '20': 'Unprocessed E-invoice', '4': 'Cancelled Invoices',
               '3': 'Transferred to GFIS', '2': 'Ready to transfer',
               '1': 'Sent for approval', '0': 'Unprocessed'}


GFIS_DATA = {}

COMBINED_DATA = {}

REQUESTED_INVOICE_STATUSES = {}

FLOW_DATA = {}


def last_column(file):
    """
    gets the index of the last column of excel data file

    :param file: path to excel file
    :return: index of the last column
    """
    gfis_excel = load_workbook(filename=f'{file}')
    gfis_sheet = gfis_excel.active
    return gfis_sheet.max_column - 1


def retrieve_gfis_data(path_wildcard):
    """

    :param path_wildcard: path to directory that contains *.xlsx files
    :return: updates the dictionary GFIS_DATA where invoice_number is a key, and
    schedule_date, parsed_date, payments are values.
    """
    for file in glob.glob(path_wildcard):
        try:
            DataFile.remove_row(file)
            gfis_excel = load_workbook(filename=f'{file}')
            gfis_sheet = gfis_excel.active
        except FileNotFoundError:
            print('File *.xlsx in <gfis> not found')
        else:
            for row in gfis_sheet.iter_rows(min_row=2, values_only=True):
                invoice_num = row[GFIS_INVOICE_COL]
                schedule_date = datetime.strftime(row[GFIS_SCHEDULE_COL], '%Y-%m-%d')
                payments = row[GFIS_PAYMENT_COL]
                spread_date = row[last_column(file)]
                parsed_date = datetime.strftime(spread_date, '%Y-%m-%d') if spread_date is not None else 'not paid'

                if invoice_num not in GFIS_DATA.keys():
                    GFIS_DATA[invoice_num] = schedule_date, parsed_date, payments
                else:
                    if payments > GFIS_DATA.get(invoice_num)[2]:
                        GFIS_DATA[invoice_num] = schedule_date, parsed_date, payments


def data_from_combined(path):
    """

    :param path: path to excel file with appended data
    :return: updates COMBINED_DATA dictionary where invoice_num is key, and status_code is value
    """
    try:
        combined_excel = load_workbook(filename=path)
        combined_sheet = combined_excel.active
    except FileNotFoundError:
        print(f'File {path} not found')
    else:
        for row in combined_sheet.iter_rows(min_row=2, values_only=True):
            invoice_num = row[BASWARE_INVOICE_COL]
            status_code = row[BASWARE_STATUS_COL]
            COMBINED_DATA[invoice_num] = status_code

def data_from_flow(path):
    """
    retrieves data from specified excel file and updates FLOW_DATA dictionary
    :param path: path to a file
    :return: updates FLOW_DATA dictionary where invoice_num is key, and approver, date_sent are values
    """
    try:
        flow_excel = load_workbook(filename=path)
        flow_sheet = flow_excel.active
    except FileNotFoundError:
        print(f'File {path} not found')
    else:
        for row in flow_sheet.iter_rows(min_row=2, values_only=True):
            invoice_num = row[FLOW_INVOICE_COL]
            approver = row[FLOW_APPROVER_COL]
            date_sent = row[FLOW_DATE_SENT_COL]
            FLOW_DATA[invoice_num] = approver, date_sent



def get_inv_status(path):
    """
    Updates dictionary where invoice number is a key and value is a string that represents actual status
    of invoice
    :param path: path to excel file that contains list of invoices to be checked
    :return: updates REQUESTED_INVOICE_STATUSES dictionary with data query
    """
    for invoice in DataFile.load_invoices(file_path=path):
        if invoice in GFIS_DATA.keys():
            REQUESTED_INVOICE_STATUSES[invoice] = \
                f'Scheduled due {GFIS_DATA[invoice][0]}, payment: {GFIS_DATA[invoice][1]}'
        elif invoice in COMBINED_DATA.keys():
            REQUESTED_INVOICE_STATUSES[invoice] = STATUS_CODES[str(COMBINED_DATA[invoice])]
        else:
            REQUESTED_INVOICE_STATUSES[invoice] = 'Missing'


def write_status(file):
    """
    Updates excel files with actual statuses of invoices
    :param file: path to excel file where statuses will be written
    :return: writes data with the status of each invoice
    """
    try:
        invoice_file = load_workbook(filename=file)
        invoice_sheet = invoice_file.active
    except FileNotFoundError:
        print(f'File {file} not found')
    else:
        for i, (k, v) in enumerate(REQUESTED_INVOICE_STATUSES.items()):
            try:
                if v == STATUS_CODES['1']:
                    invoice_sheet[f'B{i + 2}'] = f'{v} to {FLOW_DATA[k][0]} on {FLOW_DATA[k][1].split()[0]}'
                elif v == STATUS_CODES['3']:
                    invoice_sheet[f'B{i + 2}'] = f'{v}'
                    invoice_sheet[f'C{i + 2}'] = f'NO DATA IN GFIS'
                else:
                    invoice_sheet[f'B{i + 2}'] = f'{v}'
                invoice_file.save(filename=file)
            except TypeError and KeyError:
                invoice_sheet[f'B{i + 2}'] = f'{v}'
                invoice_sheet[f'C{i + 2}'] = 'Data is missing. Please check manually.'