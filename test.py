# import os
# import shutil
import pandas as pd
import glob
# import codecs
from openpyxl import load_workbook
import sys
from datetime import datetime
from dataclasses import dataclass
import json

from mappings import *



STATUS_CODES = {'R1': 'Returned', 'C1': 'Cancelled',
               '20': 'Unprocessed E-invoice', '4': 'Cancelled Invoices',
               '3': 'Transferred to GFIS', '2': 'Ready to transfer',
               '1': 'Sent for approval', '0': 'Unprocessed'}



GFIS_DATA = {}

COMBINED_DATA = {}

REQUESTED_INVOICE_STATUSES = {}

FLOW_DATA = {}



def data_from_combined(path):
    combined_excel = load_workbook(filename=path)
    combined_sheet = combined_excel.active
    for row in combined_sheet.iter_rows(min_row=2, values_only=True):
        invoice_num = row[BASWARE_INVOICE_COL]
        status_code = row[BASWARE_STATUS_COL]
        COMBINED_DATA[invoice_num] = status_code




def get_inv_status(path):
    for invoice in DataFile.load_invoices(file_path=path):
        if invoice in GFIS_DATA.keys():
            REQUESTED_INVOICE_STATUSES[invoice] = f'Scheduled due {GFIS_DATA[invoice][0]}, payment: {GFIS_DATA[invoice][1]}'
        elif invoice in COMBINED_DATA.keys():
            REQUESTED_INVOICE_STATUSES[invoice] = STATUS_CODES[str(COMBINED_DATA[invoice])]
        else:
            REQUESTED_INVOICE_STATUSES[invoice] = 'Missing'


def retrieve_gfis_data(path_wildcard):
    for file in glob.glob(path_wildcard):
        try:
            DataFile.remove_row(file)
            gfis_excel = load_workbook(filename=f'{file}')
            gfis_sheet = gfis_excel.active
        except FileNotFoundError:
            print('File *.xlsx in <gfis> not found')
        else:
            for row in gfis_sheet.iter_rows(min_row=2, values_only=True):
                invoice_num = row[GFIS_INVOICE_COL]
                schedule_date = datetime.strftime(row[GFIS_SCHEDULE_COL], '%Y-%m-%d')
                payments = row[GFIS_PAYMENT_COL]
                spread_date = row[last_column(file)]
                parsed_date = datetime.strftime(spread_date, '%Y-%m-%d') if spread_date is not None else 'not paid'

                if invoice_num not in GFIS_DATA.keys():
                    GFIS_DATA[invoice_num] = schedule_date, parsed_date, payments
                else:
                    if payments > GFIS_DATA.get(invoice_num)[2]:
                        GFIS_DATA[invoice_num] = schedule_date, parsed_date, payments






